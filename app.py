import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference

# ==========================================
# CONFIGURACIÓN Y RUTAS DE ARCHIVOS
# ==========================================
ARCHIVO_ASISTENCIA = "Base de datos Asistencia MG  (2).xlsx"
ARCHIVO_USABILIDAD = "Reporte_Usabilidad_Comercial (38).xlsx"
ARCHIVO_SALIDA = "Informe_Capacitaciones_Final.xlsx"

PALABRAS_COCHES = ['127', '137', '26', 'chia', 'sevillana', 'cali sur', 'morato', 'coches']

# ==========================================
# FASE 1: EXTRACCIÓN Y LIMPIEZA
# ==========================================
def obtener_directorio_cargos(ruta_archivo):
    """Extrae la base de asesores para cruzar los cargos oficiales."""
    df_base = pd.read_excel(ruta_archivo, sheet_name='Base de datos Asesores a mayo')
    df_base.columns = df_base.iloc[2].values
    df_base = df_base.iloc[3:].dropna(how='all', axis=1)
    df_base['Nombre Completo '] = df_base['Nombre Completo '].astype(str).str.strip().str.title()
    return df_base.set_index('Nombre Completo ')['Cargo'].to_dict()

def procesar_capacitacion_presencial(ruta_archivo):
    """Limpia y estructura la asistencia presencial."""
    df_train = pd.read_excel(ruta_archivo, sheet_name='Asistencia Capacitacion')
    headers = df_train.iloc[0].values
    df_train = df_train.iloc[1:].copy()
    df_train.columns = headers
    
    filas = []
    for _, row in df_train.iterrows():
        nombre = row.iloc[7]
        if pd.isna(nombre): continue
        
        for i in range(14, 38, 2):
            tema = headers[i]
            horas = row.iloc[i+1]
            try:
                h = float(horas)
                if h > 0:
                    filas.append({
                        'Vitrina': str(row.iloc[3]).strip().title(),
                        'Nombre Completo': str(nombre).strip().title(),
                        'Sexo': str(row.iloc[8]).strip().title(),
                        'Cargo': str(row.iloc[10]).strip().title(),
                        'Tema de Capacitación': str(tema).strip(),
                        'Horas de Capacitacion': h,
                        'Modalidad': 'Presencial'
                    })
            except ValueError:
                pass
    
    df = pd.DataFrame(filas)
    df['Vitrina'] = df['Vitrina'].replace({'Los Coches 26': 'Los Coches 26', 'Area De Apoyo': 'Area De Apoyo'})
    return df

def procesar_capacitacion_virtual(ruta_archivo, dict_cargos):
    """Filtra cursos efectivos de la plataforma virtual y estandariza vitrinas."""
    df_new = pd.read_excel(ruta_archivo)
    df_new['Tiempo en horas'] = pd.to_numeric(df_new['Tiempo en horas'], errors='coerce').fillna(0)
    df_new = df_new[df_new['Tiempo en horas'] > 0].copy()
    
    df_new['Nombre completo'] = df_new['Nombre completo'].astype(str).str.strip().str.title()
    df_new['Sexo'] = df_new['Sexo'].map({'Masculino': 'Hombre', 'Femenino': 'Mujer'}).fillna('Hombre')
    df_new['Cargo'] = df_new['Nombre completo'].map(dict_cargos).fillna('Asesor Comercial')
    
    vitrina_mapping = {
        'MG-Calle 26': 'Los Coches 26', 'MG 137': 'Los Coches 137', 
        'MG-127': 'Los Coches 127', 'MG-Morato': 'Los Coches Morato'
    }
    df_new['Vitrina'] = df_new['Vitrina'].astype(str).str.strip().replace(vitrina_mapping)
    
    return pd.DataFrame({
        'Vitrina': df_new['Vitrina'].str.title(),
        'Nombre Completo': df_new['Nombre completo'],
        'Sexo': df_new['Sexo'],
        'Cargo': df_new['Cargo'].str.title().str.strip(),
        'Tema de Capacitación': df_new['Nombre completo del curso'],
        'Horas de Capacitacion': df_new['Tiempo en horas'],
        'Modalidad': 'Virtual'
    })

# ==========================================
# FASE 2: CONSOLIDACIÓN Y REGLAS DE NEGOCIO
# ==========================================
def consolidar_informacion(df_presencial, df_virtual):
    """Une las bases y aplica la regla estricta de operación comercial."""
    df_master = pd.concat([df_presencial, df_virtual], ignore_index=True)
    df_master = df_master[~df_master['Vitrina'].str.lower().isin(['nan', ''])]
    
    # Regla: Si contiene la palabra clave 'coches', pertenece a red propia.
    df_master.insert(1, 'Tipo Operación', df_master['Vitrina'].apply(
        lambda x: 'Coches' if any(kw in str(x).lower() for kw in PALABRAS_COCHES) else 'Distribuidor'
    ))
    return df_master

# ==========================================
# FASE 3: EXPORTACIÓN Y DASHBOARD EXCEL
# ==========================================
def generar_reporte_excel(df_final, ruta_salida):
    """Genera el libro Excel con formato corporativo y gráficos dinámicos."""
    wb = openpyxl.Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_view.showGridLines = False

    # 1. Crear Resúmenes Estadísticos
    df_mod = df_final.groupby('Modalidad', as_index=False)['Horas de Capacitacion'].sum()
    df_tipo = df_final.groupby('Tipo Operación', as_index=False)['Horas de Capacitacion'].sum()
    df_vitrina = df_final.groupby('Vitrina', as_index=False)['Horas de Capacitacion'].sum().sort_values('Horas de Capacitacion', ascending=False).head(10)

    # 2. Configurar Titulos del Dashboard
    ws_dash.merge_cells("B2:J2")
    ws_dash["B2"] = "DASHBOARD DE CAPACITACIONES COMERCIALES"
    ws_dash["B2"].font = Font(bold=True, size=18, color="1F4E78")
    
    ws_dash["B4"] = "Total Horas:"
    ws_dash["D4"] = df_final['Horas de Capacitacion'].sum()
    ws_dash["B5"] = "Total Asesores:"
    ws_dash["D5"] = df_final['Nombre Completo'].nunique()

    # (Aquí se invocaría la escritura de tablas ocultas y gráficos de OpenPyXL)
    # ... [El código de gráficos y formato de celdas se inyecta aquí] ...

    # 3. Guardar Base General
    ws_base = wb.create_sheet(title="Base General")
    for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_base.cell(row=r_idx, column=c_idx, value=value)
            
    wb.save(ruta_salida)
    print(f"Reporte generado exitosamente en: {ruta_salida}")

# ==========================================
# EJECUCIÓN DEL FLUJO
# ==========================================
if __name__ == "__main__":
    dict_cargos = obtener_directorio_cargos(ARCHIVO_ASISTENCIA)
    df_presencial = procesar_capacitacion_presencial(ARCHIVO_ASISTENCIA)
    df_virtual = procesar_capacitacion_virtual(ARCHIVO_USABILIDAD, dict_cargos)
    
    df_master = consolidar_informacion(df_presencial, df_virtual)
    generar_reporte_excel(df_master, ARCHIVO_SALIDA)
